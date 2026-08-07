"""Read CC6 daily report workbooks and merge into a master table."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

DATE_RE = re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})")
HEADER_ID_HINT = "ID /"


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_date_from_filename(filename: str) -> str | None:
    """Extract YYYY-MM-DD from a daily-report filename."""
    match = DATE_RE.search(filename.replace("\xa0", " "))
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def find_header_row(ws, header_keyword: str = HEADER_ID_HINT, start_col: int = 8) -> int | None:
    keyword = header_keyword.lower()
    for row in range(1, min(ws.max_row or 1, 200) + 1):
        cell = ws.cell(row, start_col).value
        if cell is None:
            continue
        if keyword in normalize_header(cell).lower():
            return row
    return None


def read_daily_report(
    path: str | Path,
    *,
    sheet_name: str = "DAILY REPORT",
    header_keyword: str = HEADER_ID_HINT,
    data_start_col: int = 8,
    data_end_col: int = 24,
    column_name_overrides: dict[int, str] | None = None,
) -> pd.DataFrame:
    """Read the manpower/progress table (columns H-X) from a daily report.

    ``column_name_overrides`` is a ``{col_number: name}`` dict used when the
    source header is empty – useful for format migrations (e.g. new "Shift"
    column that doesn't exist in older reports).
    """
    if column_name_overrides is None:
        column_name_overrides = {24: "Shift"}
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path.name}. Sheets: {wb.sheetnames}")
        ws = wb[sheet_name]
        header_row = find_header_row(ws, header_keyword=header_keyword, start_col=data_start_col)
        if header_row is None:
            raise ValueError(f"Could not find header row containing {header_keyword!r} in {path.name}")

        headers = [
            normalize_header(ws.cell(header_row, col).value)
            or column_name_overrides.get(col, "")
            or f"Column_{col}"
            for col in range(data_start_col, data_end_col + 1)
        ]

        rows: list[list[Any]] = []
        max_row = ws.max_row or header_row
        for row in range(header_row + 1, max_row + 1):
            id_value = ws.cell(row, data_start_col).value
            if not isinstance(id_value, (int, float)):
                # Stop at first non-numeric ID after data has started; skip blank gaps before data.
                if rows:
                    break
                continue
            values = [ws.cell(row, col).value for col in range(data_start_col, data_end_col + 1)]
            rows.append(values)
    finally:
        wb.close()

    if not rows:
        return pd.DataFrame(columns=headers)

    return pd.DataFrame(rows, columns=headers)


def load_master(master_path: str | Path, sheet_name: str = "Master") -> pd.DataFrame:
    master_path = Path(master_path)
    if not master_path.exists():
        return pd.DataFrame()
    return pd.read_excel(master_path, sheet_name=sheet_name, engine="openpyxl")


def save_master(df: pd.DataFrame, master_path: str | Path, sheet_name: str = "Master") -> None:
    master_path = Path(master_path)
    master_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(master_path, sheet_name=sheet_name, index=False, engine="openpyxl")


def merge_into_master(
    report_path: str | Path,
    master_path: str | Path,
    *,
    is_update: bool = False,
    report_date: str | None = None,
    master_sheet: str = "Master",
    sheet_name: str = "DAILY REPORT",
    header_keyword: str = HEADER_ID_HINT,
    data_start_col: int = 8,
    data_end_col: int = 24,
) -> dict[str, Any]:
    """
    Merge one daily report into the master workbook.

    - Non-update: append rows.
    - Update: replace all master rows for that date, then append new rows.
    """
    report_path = Path(report_path)
    date_str = report_date or parse_date_from_filename(report_path.name)
    if not date_str:
        raise ValueError(f"Cannot parse date from filename: {report_path.name}")

    new_df = read_daily_report(
        report_path,
        sheet_name=sheet_name,
        header_keyword=header_keyword,
        data_start_col=data_start_col,
        data_end_col=data_end_col,
    )
    if new_df.empty:
        raise ValueError(f"No data rows found in {report_path.name}")

    new_df.insert(0, "Date", date_str)

    master = load_master(master_path, sheet_name=master_sheet)
    removed = 0
    if master.empty:
        merged = new_df.copy()
    else:
        # Align columns: Date + report columns; keep any unexpected master columns.
        if "Date" not in master.columns:
            master.insert(0, "Date", pd.NA)
        master["Date"] = master["Date"].astype(str)
        if is_update:
            mask = master["Date"] == date_str
            removed = int(mask.sum())
            master = master.loc[~mask].copy()
        # Union columns preserving master order then new columns
        for col in new_df.columns:
            if col not in master.columns:
                master[col] = pd.NA
        ordered_cols = list(master.columns)
        for col in new_df.columns:
            if col not in ordered_cols:
                ordered_cols.append(col)
        merged = pd.concat([master[ordered_cols], new_df.reindex(columns=ordered_cols)], ignore_index=True)

    save_master(merged, master_path, sheet_name=master_sheet)
    return {
        "date": date_str,
        "is_update": is_update,
        "rows_added": len(new_df),
        "rows_removed": removed,
        "master_rows": len(merged),
        "master_file": str(Path(master_path).resolve()),
    }
